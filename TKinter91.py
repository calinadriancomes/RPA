import tkinter 
import pandas as pd
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook  

from  openpyxl import *
from openpyxl.utils import get_column_letter


main_win = tkinter.Tk()

text0 = tkinter.Label(main_win, text = "CERERE UNICĂ DE PLATĂ din anul ",fg='black')
text0.pack()
an=tkinter.Entry(main_win)
an.insert(0,'2022')
an.pack()

text1 = tkinter.Label(main_win, text = "Selectati un documents EXcel'",fg='red')
text1.pack()

#main_win.geometry("1000x500")

#caut=str("II. A. DECLARAŢIE DE SUPRAFAŢĂ  - "+str(an)+" - IPA-Online") #Cautare antet tabel de declaratie tabele decl de supr

Shetsdecl=[]  # numar sheet care contine decl de suprafata 
ShetAdresa=[]  # nr sheet care contine date pers
ShetsdeclAnimale=[]  # numar sheet care contine date privind animalele detinute 
nrsheetgasit=0  #numarator sheet gasit decl suprafata
main_win.sourceFile = ''


coordi1=[]  #
coordi2=[]
coordi3=[]
coordi4=[]
coordi5=[]
coordi6=[]
coordi61=[] #coordinata categorie de folosinta pag 2
coordi7=[]
coordi71=[] #coordinata nr parcela pag 2 la cautare len 3 "NNa" 
coordi8=[]    #coordinate key7 nr animale
coordi9=[]  #coordinata cnp/cui
ListaCat=[]  # Lista categoriilor de folosinta
ListaCult=[] # lista culturiilor
NRrandDATA=1

TA=0

key1="Nr. Bloc fizic"
key2="Nume"
key3="Cod Pachet"
key4="Cod Siruta"
key5="Nr. Parcelă / cultură 1a; 1b; 1c ; 1d"
key6="Cod Categorie de folosinţă"
key7="47. Dacă deține animale *)"
key8="SEDIUL SOCIAL AL SOCIETĂŢII / ADRESA DE DOMICILIU: *)"
codsiruta=""
EXC='' #Variabila calcul Supr Cat de folosinta excel  
EXCTOT='='  # Variabila calcul total supr
bb=[]  #variabila lista modificare string


def chooseFile():               # Rasfoieste file
    global location, caut, coordi1, key1,key7,key9, coordi8,coordi9, Shetsdecl,nrsheetgasit,ShetsdeclAnimale,ShetAdresa, an
    location=''
    caut=str("II. A. DECLARAŢIE DE SUPRAFAŢĂ  - "+str(an.get())+" - IPA-Online") #Cautare antet tabel de declaratie tabele decl de supr
    main_win.sourceFile = filedialog.askopenfilename(parent=main_win, initialdir= "/", title='Please select a directory')
    location = main_win.sourceFile
    print(location)
    n=1
    
    #Count number of sheets
    wb = openpyxl.load_workbook(location) 
    nrSh = len(wb.sheetnames)
    print (nrSh)  
    Shetsdecl=[]
    nrsheetgasit=0
    for i in range (nrSh):        
        sh=("Table "+str(n))
        #reads sheets from excel file
        df = pd.read_excel(location,sh)
        n=n+1
        k=len(df.columns)
        for x in range (k): 
            if df.columns[x]==caut:
                        #print(df)
                Shetsdecl.append(n-1)  # adaugare numarul tabelelor la variabila Shetsdecl
                nrsheetgasit=nrsheetgasit +1
                
        ws = wb[sh]
        number_rows = ws.max_row
        number_columns = ws.max_column
        for f in range(number_columns):                 #Cautare coloana dupa nume tabel pag 1
            for s in range (number_rows):   
                coord=str(get_column_letter(f+1)+str(s+1)) #Adaugare coordonatele Celuelelor care contin nume coloana cautat la variabile coordi1...6
                if ws[coord].value==key7:
                    coordi8.append((s+1))
                    coordi8.append((f+1))
                    ShetsdeclAnimale.append(n-1)
                if ws[coord].value==key8:
                    coordi9.append((s+1))
                    coordi9.append((f+1))
                    ShetAdresa.append(n-1)
        print(location, nrsheetgasit,Shetsdecl,ShetsdeclAnimale, coordi8,ShetAdresa, coordi9, caut) 
    
#def CautTable ():           #Cautare tabele bar contin in antet 'caut' Decl Supr
                                          
                         
                
def cautColoana ():       # Cautare coloannele care contin datele necesare din tabelele gasite
    global location,TA, coordi1,coordi2,coordi3,coordi4,coordi5,coordi6,coordi61,coordi7,coordi71, key1,key2,key3,key4,key5,key6, Shetsdecl, nrsheetgasit, codsiruta,NRrandDATA, ListaCat, ListaCult    
    wb = openpyxl.open(location)
    #create new exxel sheet
    wb.create_sheet('DATA')
    wb.create_sheet('DATAA')
    coordi1=[]
    coordi2=[]
    coordi3=[]
    coordi4=[]
    coordi5=[]
    coordi6=[]
    coordi7=[]
    ListaCat=[]
    ListaCult=[]
    nrsheetgasit= len(Shetsdecl)
    NRrandDATA=1
    for i in range (nrsheetgasit):   #nrsheetgasit
        sh=("Table "+str(Shetsdecl[0]))
        ws = wb[sh]
        wd = wb['DATA']
        number_rows = ws.max_row
        number_columns = ws.max_column       
        for f in range(number_columns):                 #Cautare coloana dupa nume tabel pag 1
            for s in range (number_rows):   
                coord=str(get_column_letter(f+1)+str(s+1)) #Adaugare coordonatele Celuelelor care contin nume coloana cautat la variabile coordi1...6
                if ws[coord].value==key1:
                    coordi1.append((s+1))
                    coordi1.append((f+1))
                if ws[coord].value==key2:    
                    coordi2.append((s+1))
                    coordi2.append((f+1))
                if ws[coord].value==key3:
                    coordi3.append((s+1))
                    coordi3.append((f+1))
                if ws[coord].value==key4:
                    coordi4.append((s+1))
                    coordi4.append((f+1))
                if ws[coord].value==key5:
                    coordi5.append((s+1))
                    coordi5.append((f+1))
                if ws[coord].value==key6:
                    coordi6.append((s+1))
                    coordi6.append((f+1))
            f=f+1

        print(location, coordi1,coordi2,coordi3,coordi4,coordi5,coordi6)    
    for q in range (number_rows-coordi4[0]):                #Cod siruta pag1
        coord=get_column_letter(coordi4[1])+str(q+2+coordi4[0])
        coordD=get_column_letter(1)+str(q+2)
        wd[coordD].value=ws[coord].value
        print(coord, coordD, ws[coord].value)
        NRrandDATA=NRrandDATA+1
        
    for q in range((number_rows-coordi2[0])):               #Nume cultura pag1
        coord=get_column_letter(coordi2[1])+str(q+1+coordi2[0])
        coordD=get_column_letter(2)+str(q+1)
        wd[coordD].value=ws[coord].value
        print(coord, coordD, ws[coord].value)
        if ws[coord].value in ListaCult: 
            pass
        else:
            ListaCult.append(ws[coord].value)
            print(coord, coordD, coordi6, ListaCult)
            
    #                                                                              Pachet pag1
    for q in range((number_rows-coordi3[0])):
            coord=get_column_letter(coordi3[1])+str(q+1+coordi3[0])
            coordD=get_column_letter(3)+str(q+2)
            wd[coordD].value=ws[coord].value
            #print(coord)
            #print(ws[coord].value)
            #data3.append(str(ws[coord].value))
 

                                                                                #cod siruta pag1
    for q in range((number_rows-coordi1[0])):
            coord=get_column_letter(coordi1[1])+str(q+2+coordi1[0])
            coordD=get_column_letter(4)+str(q+2 )
            wd[coordD].value=ws[coord].value
           # print(coord)
           # print(ws[coord].value)
            #data4.append(str(ws[coord].value))
           

    codsiruta=wd[coordD].value

                                                                                #nr parc pag1
    for q in range((number_rows-1-coordi5[0])):
            coord=get_column_letter(coordi5[1])+str(q+2+coordi5[0])
            coordD=get_column_letter(5)+str(q+1)
            wd[coordD].value=ws[coord].value
          #  print(coord)
           # print(ws[coord].value)
           # data5.append(str(ws[coord].value))


                                                                        #Categorie de folosinta pag1
    for q in range((number_rows-2-coordi6[0])):
            coord=get_column_letter(coordi6[1])+str(q+3+coordi6[0])
            coordD=get_column_letter(6)+str(q+2)
            wd[coordD].value=ws[coord].value
            if ws[coord].value in ListaCat: 
                pass
            else:
                 ListaCat.append(ws[coord].value)
            print(coord, coordD, coordi6, ListaCat)

                                                                   #cod cultura pag 1
    for q in range((number_rows-1-coordi2[0])):
            coord=get_column_letter(coordi2[1]+1)+str(q+2+coordi2[0])
            coordD=get_column_letter(7)+str(q+2)
            wd[coordD].value=ws[coord].value
               
                                                                            #Suprafata pag1
    for q in range((number_rows-coordi2[0]-1)):           
            coord=get_column_letter(coordi2[1]+3)+str(q+2+coordi2[0])
            coordD=get_column_letter(8)+str(q+2)
            
            cooD=get_column_letter(10)+str(q+2)
            #gooD=get_column_letter(16)+str(q+1)
            aaa=str(ws[coord].value)
            aa=len(aaa)
            del(bb[0:(aa)])
            for w in range (aa):
                bbb=aaa[w]
                if bbb!="." and bbb!="'":                
                     bb.extend(bbb)            
        
            aaa=''.join(str(e)for e in bb)
            wd[coordD].value=aaa
            v='","'
            if len(aaa)==2:
                wd[cooD]="=(LEFT("+coordD+", 1) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -1))+0"
                print (aaa)
            if len(aaa)==3:
                wd[cooD]="=(LEFT("+coordD+", 1) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -1))+0"
                print (aaa)
            if len(aaa)==4:
                wd[cooD]="=(LEFT("+coordD+", 2) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -2))+0"
            if len(aaa)==5:
                wd[cooD]="=(LEFT("+coordD+", 3) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -3)+0)"
            if len(aaa)==6:
                wd[cooD]       
            
    if  nrsheetgasit>1:                                     #nrsheetgasit cautare in pagini cand este tabel lung mai mare de 1
        for i in range (nrsheetgasit-1):                
            sh=("Table "+str(Shetsdecl[i+1]))
            ws = wb[sh]   
            number_rows2 = ws.max_row
            number_columns2 = ws.max_column
            codsiruta=wd['A3'].value
            DCORD=NRrandDATA
            coordi7=[]
            for f in range(number_columns2):                 #Cautare coloana dupa cod siruta pag 2
                for s in range (number_rows2):   
                    coord=str(get_column_letter(f+1)+str(s+1))
                    if ws[coord].value==codsiruta:
                        coordi7.append((s+1))
                        coordi7.append((f+1))
                    else: exit
            
            for i in range(number_columns2):            # cautare loc coloana nr parcela
                for k in range(number_rows2):
                    coord=str(get_column_letter(i+1)+str(k+1))
                    aaa=str(ws[coord].value)
                    aa=len(aaa)
                    if aa==3:
                        if  aaa[2]=='a' :
                            if aaa[2]!= '0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9' :
                                if aaa[0]=='0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9' :
                                    if aaa[1]=='0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9':
                                        coordi71.append((k+1))
                                        coordi71.append((i+1))         
                                        print(coord)
                    if aa==4:
                        if  aaa[3]=='a' :
                            if aaa[3]!= '0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9' :
                                if aaa[0]=='0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9' :
                                    if aaa[1]=='0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9':
                                        if aaa[2]=='0' or '1' or '2' or '3' or '4' or '5' or '6' or '7' or '8' or '9':
                                            coordi71.append((k+1))
                                            coordi71.append((i+1))   
                    if aaa=='PP':
                        coordi61.append((k+1))
                        coordi61.append((i+1)) 
                    if aaa=='TA':
                        coordi61.append((k+1))
                        coordi61.append((i+1))
                    if aaa=='CP':
                        coordi61.append((k+1))
                        coordi61.append((i+1))




            
            for q in range (int((len(coordi7))/2)):                                      #Cod siruta pag2
                coord=get_column_letter(coordi7[1])+str(q+3)
                coordD=get_column_letter(1)+str(DCORD)
                wd[coordD].value=ws[coord].value
                DCORD=DCORD+1
            DCORD=NRrandDATA    
            for q in range (int((len(coordi7))/2)):                                     #Nume cultura pag2   
                coord=get_column_letter(coordi71[1]+1)+str(q+3)                            
                coordD=get_column_letter(2)+str(str(DCORD))
                wd[coordD].value=ws[coord].value  
                DCORD=DCORD+1
                if ws[coord].value in ListaCult: 
                    pass
                else:
                    ListaCult.append(ws[coord].value)
            print(coord, coordD, coordi6, ListaCult)
            DCORD=NRrandDATA 
            
            for q in range (int((len(coordi7))/2)):                                   # cod pachet  pag2  
                coord=get_column_letter(coordi7[1]+8)+str(q+3)                            
                coordD=get_column_letter(3)+str(DCORD)
                wd[coordD].value=ws[coord].value   
                DCORD=DCORD+1                
                #print (coordi7[0], ws[coord].value, wd[coordD].value, coord,coordD )  
            DCORD=NRrandDATA  
            
            for q in range (int((len(coordi7))/2)):                                   # bloc fizic  pag2  
                coord=get_column_letter(coordi7[1]+1)+str(q+3)                            
                coordD=get_column_letter(4)+str(DCORD)
                wd[coordD].value=ws[coord].value 
                DCORD=DCORD+1
            DCORD=NRrandDATA  
            
            for q in range (int((len(coordi7))/2)):                                   # nr parc  pag2  
                coord=get_column_letter(coordi71[1])+str(q+3)                            
                coordD=get_column_letter(5)+str(DCORD)
                wd[coordD].value=ws[coord].value  
                DCORD=DCORD+1
            DCORD=NRrandDATA 

            for q in range (int((len(coordi7))/2)):                                   # categorie de folosinta  pag2  
                coord=get_column_letter(coordi61[1])+str(q+3)                            
                coordD=get_column_letter(6)+str(DCORD)
                wd[coordD].value=ws[coord].value  
                DCORD=DCORD+1
                if ws[coord].value in ListaCat: 
                    pass
                else:
                    ListaCat.append(ws[coord].value)
                print(coord, coordD, coordi6, ListaCat)
            DCORD=NRrandDATA 
                      
            for q in range (int((len(coordi7))/2)):                                   # cod cultura pag2  
                coord=get_column_letter(coordi71[1]+2)+str(q+3)                            
                coordD=get_column_letter(7)+str(DCORD)
                wd[coordD].value=ws[coord].value  
                DCORD=DCORD+1
            DCORD=NRrandDATA                 
            
            for q in range (int((len( coordi7))/2)):                                   # suprafata parc  pag2  
                coord=get_column_letter(coordi71[1]+4)+str(q+3)                            
                coordD=get_column_letter(8)+str(DCORD)
                wd[coordD].value=ws[coord].value  
                cooD=get_column_letter(10)+str(DCORD)          #loc coloana cu ,
                #gooD=get_column_letter(16)+str(q+1)
                aaa=str(ws[coord].value)
                aa=len(aaa)
                del(bb[0:(aa)])
                for w in range (aa):
                    bbb=aaa[w]
                    if bbb!="." and bbb!="'":                
                     bb.extend(bbb)            
        
                aaa=''.join(str(e)for e in bb)
                wd[coordD].value=aaa
                v='","'
                if len(aaa)==2:
                    wd[cooD]="=(LEFT("+coordD+", 1) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -1))+0"
                    print (aaa)
                if len(aaa)==3:
                    wd[cooD]="=(LEFT("+coordD+", 1) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -1))+0"
                    print (aaa)
                if len(aaa)==4:
                    wd[cooD]="=(LEFT("+coordD+", 2) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -2))+0"
                if len(aaa)==5:
                    wd[cooD]="=(LEFT("+coordD+", 3) &"+ v +"& RIGHT(" +coordD+ ", LEN("+coordD+") -3))+0"
                if len(aaa)==6:
                    wd[cooD]      
                DCORD=DCORD+1
            DCORD=NRrandDATA    
                #print (coordi7[0], ws[coord].value, wd[coordD].value, coord,coordD )  



                
            NRrandDATA= NRrandDATA+int((len(coordi7))/2)      
            
    
        
        
        print ( nrsheetgasit)
            
        print(NRrandDATA, coordi1,coordi2,coordi3,coordi4,coordi5,coordi6)    
        print(coordi1[0],coordi2[0],coordi3[0],coordi4[0],coordi5[0],coordi6[0])   
        print(number_columns,number_rows, location, Shetsdecl, nrsheetgasit, ListaCat, ListaCult)      


     #bloc fizic pag1
            
           # for i in range(number_rows-coordi1[0]):
            #    coord=get_column_letter(coordi1[0])+str(q+1+coordi1[1])
             #   coordD=get_column_letter(1)+str(q+1)
                #wd[coordD].value=ws[coord].value
                #print(coord)
                #print(ws[coord].value)
                #data1.append(ws[coord].value)
        
        
        
     # salvare tabel
    wb.save(location)   
        
        #print (ws, sh,DATA,number_rows,number_columns)
        #wd = wb["DATA"]
        
def calcul():
    global location, EXC,EXCTOT, NRrandDATA, ListaCat, ListaCult,codsiruta 
    wb = openpyxl.open(location)
    wd = wb ['DATA']
    gasit=0
    codsiruta= wd['A5'].value
    number_rows2 = wd.max_row
    number_columns2 = wd.max_column
      
        
    
    for i in range (len(ListaCat)):
        CF=ListaCat[i-1]
        EXC=''
        gasit=1
        for p in range (NRrandDATA): 
            #EXCTOT='='
            coord=get_column_letter(6)+str(p+2)
            coords=get_column_letter(10)+str(p+2)
            COORD=get_column_letter(15)+str(5+i)
            COORDN=get_column_letter(16)+str(5+i)
            #EXCTOT=EXCTOT+"+"+str(coords)          
                             
            if wd[coord].value==CF and gasit ==1:
                EXC=EXC+"="
                gasit=2
                
            if wd[coord].value==CF and gasit==2:
                EXC=EXC+"+"+str(coords)
            #if wd[coord].value != ' ':     
            #EXCTOT=EXCTOT+"+"+str(coords)
             
           
        wd[COORD].value=ListaCat[i-1]
        wd[COORDN].value=EXC
        wd['L6'].value='Suprafata totala' 
        #wd['M6'].value=EXCTOT
        print ( EXC, coordi7)
    
    
    for i in range (len(ListaCult)):
        CF=ListaCult[i-1]
        EXC=''
        gasit=1
        for p in range (NRrandDATA): 
            
            coord=get_column_letter(2)+str(p+2)
            coords=get_column_letter(10)+str(p+2)
            COORD=get_column_letter(12)+str(8+i)
            COORDN=get_column_letter(13)+str(8+i)
                   
                             
            if wd[coord].value==CF and gasit ==1:
                EXC=EXC+"="
                gasit=2
                
            if wd[coord].value==CF and gasit==2:
                EXC=EXC+"+"+str(coords)
            
             
           
        wd[COORD].value=ListaCult[i-1]
        wd[COORDN].value=EXC
        
        #wd['M6'].value=EXCTOT
    print(NRrandDATA)    
     # salvare tabel
    wb.save(location)                 
                    
           





b_chooseFile = tkinter.Button(main_win, text = "Răsfoiește", width = 10, height = 1, command = chooseFile)
b_chooseFile.pack()
b_chooseTab = tkinter.Button(main_win, text = "Exportă datele", width = 10, height = 1, command = cautColoana)
b_chooseTab.pack()
b_chooseTab = tkinter.Button(main_win, text = "Calcul", width = 10, height = 1, command = calcul)
b_chooseTab.pack()
b_exit = tkinter.Button(main_win, text = "Exit",command = main_win.destroy)
b_exit.pack()
print(main_win.sourceFile )
main_win.mainloop()




